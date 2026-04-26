"""
Multi-source Talent Sourcing Portal — Streamlit app

Search sources  : Exa · Google CSE · Apollo · People Data Labs · Crustdata
Enrichment      : FullEnrich · Lusha
Scoring         : Rule-based weighted rubric (100 pts)

Deploy on Streamlit Cloud:
  1. Push to GitHub → share.streamlit.io → New app → app.py
  2. Advanced settings → Secrets — add whichever keys you have:

     APP_PASSWORD        = "your-team-password"
     EXA_API_KEY         = "..."          # exa.ai
     GOOGLE_API_KEY      = "..."          # console.cloud.google.com
     GOOGLE_CSE_ID       = "..."          # programmablesearchengine.google.com
     APOLLO_API_KEY      = "..."          # app.apollo.io
     PDL_API_KEY         = "..."          # peopledatalabs.com
     CRUSTDATA_API_KEY   = "..."          # crustdata.com
     FULLENRICH_API_KEY  = "..."          # fullenrich.com  (enrichment)
     LUSHA_API_KEY       = "..."          # lusha.com       (enrichment)

  At least one search source key is required. All others are optional —
  the app gracefully skips any source without a key.
"""

import re
import urllib.parse
from datetime import date
from io import BytesIO

import pandas as pd
import requests
import streamlit as st

# ── Page config ───────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Talent Sourcing Portal",
    page_icon="🔍",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Auth gate ─────────────────────────────────────────────────────────────────

def check_password():
    try:
        correct = st.secrets["APP_PASSWORD"]
    except Exception:
        correct = None
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if not st.session_state.authenticated:
        st.markdown("## 🔐 Talent Sourcing Portal")
        st.markdown("Enter the team password to continue.")
        col1, _ = st.columns([2, 3])
        with col1:
            entered = st.text_input("Password", type="password", key="pw_input")
            if st.button("Login", type="primary"):
                if correct and entered == correct:
                    st.session_state.authenticated = True
                    st.rerun()
                elif not correct and entered:
                    st.session_state.authenticated = True
                    st.rerun()
                else:
                    st.error("Incorrect password.")
        st.stop()

check_password()

# ── Secrets helper ────────────────────────────────────────────────────────────

def get_secret(key: str) -> str:
    try:
        raw = st.secrets[key]
    except Exception:
        return ""
    return str(raw).strip().encode("ascii", errors="ignore").decode("ascii")

# ── JD intelligence — language-agnostic role/industry detection ───────────────

SENIORITY_WORDS = {
    "partner": 5, "vice president": 5, "vp": 5,
    "director": 4, "principal": 4, "geschäftsführer": 4, "direktor": 4,
    "manager": 3, "staff": 3, "lead": 3, "leiter": 3, "projektleiter": 3, "teamleiter": 3,
    "senior": 2, "leitend": 2,
    "associate": 1,
    "analyst": 0, "junior": 0, "intern": 0, "praktikant": 0,
}

# Role archetypes: (english search terms, detection keywords DE+EN)
ROLE_ARCHETYPES = [
    (["sales manager", "account manager", "business development manager"],
     ["vertrieb", "verkauf", "sales", "account manager", "key account", "vertriebsmanager"]),
    (["strategy consultant", "management consultant"],
     ["strategie", "strategy consultant", "management consultant", "unternehmensberater", "beratung"]),
    (["software engineer", "developer", "software developer"],
     ["entwickler", "software engineer", "developer", "programmier", "coding", "python", "java"]),
    (["product manager"],
     ["produktmanager", "product manager", "produktmanagement"]),
    (["project manager", "program manager"],
     ["projektmanager", "projektleiter", "project manager", "program manager", "pmo"]),
    (["data scientist", "data analyst"],
     ["data scientist", "data analyst", "machine learning", "ki", "artificial intelligence"]),
    (["marketing manager"],
     ["marketing", "marketingmanager", "brand manager"]),
    (["hr manager", "talent acquisition", "recruiter"],
     ["personalreferent", "hr manager", "human resources", "recruiting", "talent"]),
    (["finance manager", "controller"],
     ["controller", "finance manager", "finanzmanager", "buchhaltung", "accounting"]),
    (["operations manager"],
     ["operations", "betrieb", "supply chain", "logistik", "logistics"]),
    (["digital transformation manager"],
     ["digitalisierung", "digital transformation", "digitalen", "digital portfolio"]),
]

# Industry archetypes: (english search terms, detection keywords DE+EN)
INDUSTRY_ARCHETYPES = [
    (["healthcare", "medical technology", "medtech", "hospital"],
     ["gesundheit", "medizin", "medizintechnik", "healthineers", "klinik", "krankenhaus",
      "healthcare", "medical", "radiolog", "diagnostik", "patient"]),
    (["pharmaceutical", "biotech", "life sciences"],
     ["pharma", "pharmaceutical", "biotech", "arzneimittel", "wirkstoff"]),
    (["automotive"],
     ["automobil", "automotive", "fahrzeug", "pkw", "zulieferer"]),
    (["energy", "renewables", "utilities"],
     ["energie", "energy", "strom", "solar", "wind", "erneuerbar", "utilities"]),
    (["financial services", "banking", "insurance"],
     ["finanz", "bank", "versicherung", "finance", "insurance", "kapital"]),
    (["technology", "software", "IT"],
     ["software", "technologie", "technology", "saas", "cloud", "plattform"]),
    (["industrial", "manufacturing", "engineering"],
     ["industrie", "industrial", "fertigung", "maschinenbau", "manufacturing", "engineering"]),
    (["consulting", "professional services"],
     ["beratung", "consulting", "consultant", "advisory"]),
    (["retail", "consumer goods", "FMCG"],
     ["handel", "retail", "konsumgüter", "fmcg", "consumer"]),
]

DE_STOPWORDS = {
    "sie", "und", "der", "die", "das", "in", "für", "mit", "von", "auf", "an",
    "zu", "den", "dem", "ein", "eine", "einer", "eines", "ist", "sind", "haben",
    "ihre", "ihren", "ihrem", "auch", "als", "oder", "aber", "wie", "bei", "aus",
    "nach", "über", "unter", "durch", "werden", "wird", "kann", "können", "unser",
    "unsere", "dieser", "diese", "dieses", "sowie", "ggf", "insbesondere", "dabei",
    "gemeinsam", "beim", "diesem", "dieser", "werden", "unseren", "einen", "neue",
    "neue", "neuen", "weitere", "weiteren", "mehr", "sehr", "alle", "wir",
}
EN_STOPWORDS = {
    "the", "and", "for", "with", "from", "that", "this", "you", "your", "our",
    "are", "will", "have", "has", "been", "able", "also", "their", "they", "can",
    "should", "must", "within", "across", "through", "between", "into", "about",
    "including", "such", "both", "each", "its", "not", "all", "more", "new",
}

def extract_jd_key_terms(jd_text: str, top_n: int = 25) -> list:
    """Extract the most distinctive content words from the JD for dynamic scoring."""
    from collections import Counter
    text = jd_text.lower()
    # Extract words 4+ chars, including German umlauts
    words = re.findall(r'\b[a-zäöüß]{4,}\b', text)
    stopwords = DE_STOPWORDS | EN_STOPWORDS
    freq = Counter(w for w in words if w not in stopwords)
    return [w for w, _ in freq.most_common(top_n)]


def detect_role_and_industry(jd_text: str):
    """Detect the best-matching role archetype and industries from JD text."""
    text = jd_text.lower()

    # Score each role archetype
    role_scores = []
    for search_terms, keywords in ROLE_ARCHETYPES:
        score = sum(1 for kw in keywords if kw in text)
        role_scores.append((score, search_terms))
    role_scores.sort(reverse=True)
    # Pick top role (fall back to a generic professional if nothing matches)
    best_role_score, best_role_terms = role_scores[0]
    role_titles = best_role_terms if best_role_score > 0 else ["professional", "specialist", "manager"]

    # Score each industry archetype — can match multiple
    industry_search_terms = []
    for search_terms, keywords in INDUSTRY_ARCHETYPES:
        if any(kw in text for kw in keywords):
            industry_search_terms.extend(search_terms[:2])
    if not industry_search_terms:
        industry_search_terms = ["technology", "business"]

    return role_titles, industry_search_terms[:4]


# ── JD parser ─────────────────────────────────────────────────────────────────

def parse_jd(jd_text: str, location_filter: str, seniority_opt: str,
             industry_tags: list, extra_must: str) -> dict:
    """Return structured requirements dict from JD text + UI filters."""
    text = jd_text.lower()

    # Detect role and industry from JD content (not hardcoded defaults)
    role_titles, detected_industries = detect_role_and_industry(jd_text)

    seniority_level_map = {
        "Senior / Manager": 2, "Staff / Principal": 3, "Director / VP": 4, "Any": 2,
    }
    seniority_label_map = {
        "Senior / Manager": "senior", "Staff / Principal": "principal",
        "Director / VP": "director", "Any": "",
    }
    seniority_level = seniority_level_map.get(seniority_opt, 2)
    seniority_label = seniority_label_map.get(seniority_opt, "senior")
    # Override from JD text
    if any(w in text for w in ["director", "vp", "vice president", "direktor"]):
        seniority_label = "director"
    elif any(w in text for w in ["principal", "staff", "leitend"]):
        seniority_label = "principal"
    elif any(w in text for w in ["senior", "sr.", "erfahren"]):
        seniority_label = "senior"

    locations = [l.strip() for l in location_filter.split(",") if l.strip()] or ["Germany"]
    location_country = locations[0].lower()

    # Language signals
    lang_signals = []
    if any(w in text for w in ["german", "deutsch", "deutschkenntnisse", "muttersprache"]):
        lang_signals.append("German")
    if any(w in text for w in ["english", "englisch", "englishkenntnisse"]):
        lang_signals.append("English")

    # Extract company name if mentioned prominently (e.g. "Siemens Healthineers")
    company_signals = []
    for company in ["siemens healthineers", "siemens", "bmw", "volkswagen", "bosch",
                    "bayer", "basf", "allianz", "sap", "deutsche bank", "lufthansa"]:
        if company in text:
            company_signals.append(company.title())
            break

    # Build the search query from detected role + industry + context
    query_parts = []
    if seniority_label:
        query_parts.append(seniority_label)
    query_parts.extend(role_titles[:2])
    query_parts.extend(detected_industries[:2])
    query_parts.extend(lang_signals)
    query_parts.append(locations[0])
    if company_signals:
        query_parts.extend(company_signals)
    if extra_must:
        query_parts.append(extra_must)

    # Key terms for dynamic scoring
    key_terms = extract_jd_key_terms(jd_text)

    return {
        "role_titles": role_titles,
        "seniority_level": seniority_level,
        "seniority_label": seniority_label,
        "locations": locations,
        "location_country": location_country,
        "industries": industry_tags or detected_industries,
        "detected_industries": detected_industries,
        "years_exp_min": 3,
        "query_text": " ".join(dict.fromkeys(query_parts)),  # dedup preserving order
        "jd_key_terms": key_terms,
    }

# ── Candidate normalisation ───────────────────────────────────────────────────

def norm(overrides: dict) -> dict:
    base = {
        "name": "", "headline": "", "company_name": "", "location": "",
        "linkedin_url": "", "profile_text": "", "source": "", "email": "", "phone": "",
    }
    base.update(overrides)
    return base

# ── Source: Exa ───────────────────────────────────────────────────────────────

def search_exa(query: str, api_key: str, n: int) -> list:
    resp = requests.post(
        "https://api.exa.ai/search",
        headers={"x-api-key": api_key, "Content-Type": "application/json"},
        json={
            "query": query,
            "category": "people",
            "type": "auto",
            "num_results": n,
            "contents": {"text": {"max_characters": 8000}},
        },
        timeout=30,
    )
    resp.raise_for_status()
    out = []
    for r in resp.json().get("results", []):
        url = r.get("url", "")
        text = r.get("text") or ""
        title = r.get("title", "")
        out.append(norm({
            "name": r.get("author") or title.split(" - ")[0].split("|")[0].strip(),
            "headline": title,
            "linkedin_url": url if "linkedin.com/in/" in url else "",
            "profile_text": text[:6000],
            "source": "Exa",
        }))
    return out

# ── Source: Google CSE (X-Ray LinkedIn) ──────────────────────────────────────

def search_google_cse(query: str, api_key: str, cse_id: str, n: int) -> list:
    xray = f'site:linkedin.com/in/ {query}'
    resp = requests.get(
        "https://www.googleapis.com/customsearch/v1",
        params={"key": api_key, "cx": cse_id, "q": xray, "num": int(min(n, 10))},
        timeout=30,
    )
    resp.raise_for_status()
    out = []
    for item in resp.json().get("items", []):
        title = item.get("title", "")
        name = title.split(" - ")[0].split("|")[0].strip()
        out.append(norm({
            "name": name,
            "headline": title,
            "linkedin_url": item.get("link", ""),
            "profile_text": item.get("snippet", ""),
            "source": "Google CSE",
        }))
    return out

# ── Source: Apollo ────────────────────────────────────────────────────────────

def search_apollo(jd_reqs: dict, api_key: str, n: int) -> list:
    resp = requests.post(
        "https://api.apollo.io/api/v1/mixed_people_search",
        headers={
            "Content-Type": "application/json",
            "Cache-Control": "no-cache",
            "x-api-key": api_key,
        },
        json={
            "person_titles": jd_reqs["role_titles"],
            "person_locations": jd_reqs["locations"],
            "page": 1,
            "per_page": min(n, 25),
        },
        timeout=30,
    )
    resp.raise_for_status()
    out = []
    for p in resp.json().get("people", []):
        org = p.get("organization") or {}
        history = p.get("employment_history") or []
        parts = [p.get("title", ""), org.get("name", ""), p.get("city", ""), p.get("country", "")]
        for job in history[:5]:
            t = job.get("title", "")
            co = job.get("organization_name", "")
            if t or co:
                parts.append(f"{t} at {co}".strip(" at"))
        out.append(norm({
            "name": f"{p.get('first_name', '')} {p.get('last_name', '')}".strip(),
            "headline": p.get("title", ""),
            "company_name": org.get("name", ""),
            "location": f"{p.get('city', '')} {p.get('country', '')}".strip(),
            "linkedin_url": p.get("linkedin_url") or "",
            "email": p.get("email") or "",
            "profile_text": " | ".join(filter(None, parts)),
            "source": "Apollo",
        }))
    return out

# ── Source: People Data Labs ──────────────────────────────────────────────────

def search_pdl(jd_reqs: dict, api_key: str, n: int) -> list:
    titles_clause = " OR ".join(
        f"job_title LIKE '%{t}%'" for t in jd_reqs["role_titles"][:3]
    )
    country = jd_reqs["location_country"]
    sql = (
        f"SELECT * FROM person WHERE ({titles_clause}) "
        f"AND location_country='{country}' LIMIT {n}"
    )
    resp = requests.get(
        "https://api.peopledatalabs.com/v5/person/search",
        headers={"X-Api-Key": api_key},
        params={"sql": sql, "size": n},
        timeout=30,
    )
    resp.raise_for_status()
    out = []
    for p in resp.json().get("data", []):
        experience = p.get("experience") or []
        parts = [p.get("job_title", ""), p.get("job_company_name", ""), p.get("location_name", "")]
        for exp in experience[:5]:
            t = (exp.get("title") or {}).get("name", "")
            co = (exp.get("company") or {}).get("name", "")
            if t or co:
                parts.append(f"{t} at {co}".strip(" at"))
        skills = [s.get("name", "") for s in (p.get("skills") or [])]
        if skills:
            parts.append("Skills: " + ", ".join(skills[:10]))
        li_id = p.get("linkedin_id", "")
        li_url = p.get("linkedin_url") or (f"https://linkedin.com/in/{li_id}" if li_id else "")
        emails = p.get("emails") or [{}]
        out.append(norm({
            "name": p.get("full_name", ""),
            "headline": p.get("job_title", ""),
            "company_name": p.get("job_company_name", ""),
            "location": p.get("location_name", ""),
            "linkedin_url": li_url,
            "email": emails[0].get("address", ""),
            "profile_text": " | ".join(filter(None, parts)),
            "source": "PDL",
        }))
    return out

# ── Source: Crustdata ─────────────────────────────────────────────────────────

def search_crustdata(jd_reqs: dict, api_key: str, n: int) -> list:
    resp = requests.post(
        "https://api.crustdata.com/screener/person/search",
        headers={"Authorization": f"Token {api_key}", "Content-Type": "application/json"},
        json={
            "filters": {
                "job_title": jd_reqs["role_titles"][0],
                "location": jd_reqs["locations"][0],
            },
            "page": 1,
            "limit": n,
        },
        timeout=30,
    )
    resp.raise_for_status()
    data = resp.json()
    people = data.get("data") or data.get("profiles") or data.get("results") or []
    out = []
    for p in people:
        out.append(norm({
            "name": p.get("full_name") or f"{p.get('first_name', '')} {p.get('last_name', '')}".strip(),
            "headline": p.get("headline") or p.get("title", ""),
            "company_name": p.get("company") or p.get("current_company", ""),
            "location": p.get("location", ""),
            "linkedin_url": p.get("linkedin_url", ""),
            "profile_text": p.get("summary") or p.get("headline", ""),
            "source": "Crustdata",
        }))
    return out

# ── Enrichment: FullEnrich ────────────────────────────────────────────────────

def enrich_fullenrich(candidates: list, api_key: str) -> list:
    enriched = []
    for c in candidates:
        if not c.get("linkedin_url"):
            enriched.append(c)
            continue
        try:
            resp = requests.post(
                "https://api.fullenrich.com/v1/enrich/person",
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json={"linkedin_url": c["linkedin_url"]},
                timeout=15,
            )
            if resp.ok:
                d = resp.json().get("person") or {}
                if d.get("full_name"):
                    c["name"] = d["full_name"]
                if d.get("headline"):
                    c["headline"] = d["headline"]
                if d.get("email"):
                    c["email"] = d["email"]
                if d.get("summary"):
                    c["profile_text"] = (c["profile_text"] + " " + d["summary"]).strip()
                c["source"] += " +FullEnrich"
        except Exception:
            pass
        enriched.append(c)
    return enriched

# ── Enrichment: Lusha ─────────────────────────────────────────────────────────

def enrich_lusha(candidates: list, api_key: str) -> list:
    enriched = []
    for c in candidates:
        parts = (c.get("name") or "").split()
        if len(parts) < 2 or not c.get("company_name"):
            enriched.append(c)
            continue
        try:
            resp = requests.get(
                "https://api.lusha.com/person",
                headers={"api_key": api_key},
                params={
                    "firstName": parts[0],
                    "lastName": " ".join(parts[1:]),
                    "company": c["company_name"],
                },
                timeout=15,
            )
            if resp.ok:
                d = resp.json()
                emails = d.get("emailAddresses") or []
                phones = d.get("phoneNumbers") or []
                if emails:
                    c["email"] = emails[0].get("emailAddress", c.get("email", ""))
                if phones:
                    c["phone"] = phones[0].get("localizedPhoneNumber", "")
                c["source"] += " +Lusha"
        except Exception:
            pass
        enriched.append(c)
    return enriched

# ── Merge & Deduplicate ───────────────────────────────────────────────────────

def merge_candidates(candidate_lists: list) -> list:
    all_cands = [c for lst in candidate_lists for c in lst]
    seen_li, seen_names, merged = set(), set(), []
    for c in all_cands:
        li = (c.get("linkedin_url") or "").rstrip("/").lower()
        name = (c.get("name") or "").lower().strip()
        if li and li in seen_li:
            # Merge into existing record
            for ex in merged:
                if (ex.get("linkedin_url") or "").rstrip("/").lower() == li:
                    if c["source"] not in ex["source"]:
                        ex["source"] += f", {c['source']}"
                    if len(c.get("profile_text", "")) > len(ex.get("profile_text", "")):
                        ex["profile_text"] = c["profile_text"]
                    if not ex.get("email") and c.get("email"):
                        ex["email"] = c["email"]
                    if not ex.get("company_name") and c.get("company_name"):
                        ex["company_name"] = c["company_name"]
                    break
            continue
        if name and name in seen_names and not li:
            continue
        if li:
            seen_li.add(li)
        if name:
            seen_names.add(name)
        merged.append(c)
    return merged

# ── Scoring ───────────────────────────────────────────────────────────────────

def score_candidate(profile_text: str, jd_reqs: dict) -> dict:
    """
    Score a candidate against the JD. Scoring is fully dynamic — driven by
    key terms extracted from the actual JD text, not hardcoded role assumptions.
    """
    text = (profile_text or "").lower()

    # 1. JD key-term overlap (40 pts) — how many distinctive JD terms appear in profile
    key_terms = jd_reqs.get("jd_key_terms", [])
    if key_terms:
        hits = {term: (term in text) for term in key_terms}
        hit_count = sum(hits.values())
        # Generous curve: hitting 40% of terms = full marks
        term_score = min(40, round(40 * hit_count / max(len(key_terms) * 0.4, 1)))
        term_detail = {k: v for k, v in hits.items() if v}  # only show hits
    else:
        term_score = 20  # no terms extracted — neutral
        term_detail = {}

    # 2. Role match (20 pts) — do role titles from JD appear in profile?
    role_titles = jd_reqs.get("role_titles", [])
    role_hits = sum(1 for r in role_titles if r.lower() in text)
    role_score = min(20, round(20 * role_hits / max(len(role_titles), 1)))

    # 3. Seniority (15 pts)
    target = jd_reqs.get("seniority_level", 2)
    detected = max((level for word, level in SENIORITY_WORDS.items() if word in text), default=0)
    diff = abs(detected - target)
    seniority_score = 15 if diff == 0 else (8 if diff == 1 else 0)

    # 4. Industry match (15 pts) — do detected industry terms appear?
    detected_industries = jd_reqs.get("detected_industries", [])
    if detected_industries and any(ind.lower() in text for ind in detected_industries):
        industry_score = 15
    elif detected_industries and any(
        word in text for ind in detected_industries for word in ind.split()
    ):
        industry_score = 8
    else:
        industry_score = 0

    # 5. Location (5 pts)
    locs = [l.lower() for l in jd_reqs.get("locations", ["germany"])]
    location_signal = 5 if any(l in text for l in locs) else (2 if "remote" in text else 0)

    # Experience years (bonus, up to 5 pts)
    years = [int(y) for y in re.findall(r'(\d+)\s*(?:year|yr|jahre|jahren)', text)]
    max_years = max(years, default=0)
    min_exp = jd_reqs.get("years_exp_min", 3)
    exp_bonus = 5 if max_years >= min_exp else (3 if max_years >= min_exp - 1 else 0)

    # Data quality penalty
    wc = len(text.split())
    quality_penalty = 15 if wc < 30 else (8 if wc < 100 else 0)
    quality = "sparse" if wc < 30 else ("partial" if wc < 100 else "full")

    total = term_score + role_score + seniority_score + industry_score + location_signal + exp_bonus - quality_penalty

    return {
        "must_have_detail": term_detail,   # reuse this field for display
        "must_have": term_score,
        "role_match": role_score,
        "seniority": seniority_score,
        "industry": industry_score,
        "location_signal": location_signal,
        "experience": exp_bonus,
        "company_signal": 0,
        "quality_penalty": quality_penalty,
        "data_quality": quality,
        "total": max(0, min(100, total)),
    }


def grade_label(score: int) -> str:
    if score >= 75:
        return "🟢 Strong Match"
    if score >= 50:
        return "🟡 Potential"
    return "🔴 Weak Match"

# ── Boolean string builder ────────────────────────────────────────────────────

def build_boolean_string(jd_text: str, refinement: str = "") -> str:
    """Build a Google X-Ray Boolean string from JD content — role agnostic."""
    ref = refinement.lower()

    # Detect role terms from JD
    role_titles, detected_industries = detect_role_and_industry(jd_text)
    role_quoted = " OR ".join(f'"{t}"' for t in role_titles[:3])

    # Seniority
    text = jd_text.lower()
    seniority_terms = []
    if any(w in text for w in ["senior", "sr.", "leitend", "erfahren"]):
        seniority_terms.append('"senior"')
    if any(w in text for w in ["manager", "leiter", "lead"]):
        seniority_terms.append('"manager"')
    if any(w in text for w in ["director", "direktor", "vp"]):
        seniority_terms.append('"director"')
    if not seniority_terms:
        seniority_terms = ['"senior"', '"manager"']
    seniority_str = f'({" OR ".join(seniority_terms)})'

    # Language filter
    lang_filter = ""
    if any(w in text for w in ["deutsch", "german", "muttersprache"]):
        lang_filter = '("German" OR "Deutsch")'

    # Refinement overrides
    company_filter = ""
    if refinement:
        ref_words = ref.split()
        # Extract quoted company or focus hints
        company_filter = refinement[:80]  # pass through as-is for specificity

    # Exclusions
    exclude = "-recruiter -intern -junior -praktikant"
    if any(w in ref for w in ["no automotive", "remove automotive"]):
        exclude += " -automotive"

    parts = [
        'site:linkedin.com/in/',
        f'({role_quoted})',
        seniority_str,
    ]
    if lang_filter:
        parts.append(lang_filter)
    if company_filter:
        parts.append(f'"{company_filter}"')
    parts.append(exclude)
    return " ".join(parts)

# ── Excel export ──────────────────────────────────────────────────────────────

def build_excel_bytes(df: pd.DataFrame, jd_text: str, boolean_str: str) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates"

        headers = ["Rank", "Score", "Grade", "Name", "Headline", "Company",
                   "Location", "LinkedIn URL", "Email", "Phone", "Source",
                   "Must-Have Skills", "Data Quality", "Profile Highlights"]
        col_widths = [5, 7, 18, 25, 35, 25, 20, 30, 28, 16, 20, 45, 12, 50]

        hfill = PatternFill("solid", fgColor="1F3864")
        hfont = Font(bold=True, color="FFFFFF", size=11)
        green  = PatternFill("solid", fgColor="C6EFCE")
        yellow = PatternFill("solid", fgColor="FFEB9C")
        red    = PatternFill("solid", fgColor="FFC7CE")
        wrap   = Alignment(wrap_text=True, vertical="top")
        center = Alignment(horizontal="center", vertical="center")

        ws.append(headers)
        for i, (cell, w) in enumerate(zip(ws[1], col_widths), 1):
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = center
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for _, row in df.iterrows():
            score = int(row.get("score", 0))
            fill = green if score >= 75 else (yellow if score >= 50 else red)
            mh = row.get("must_have_detail", {})
            mh_str = ", ".join(f"{k} {'✓' if v else '✗'}" for k, v in mh.items()) if isinstance(mh, dict) else ""
            ws.append([
                int(row.get("rank", 0)),
                score,
                grade_label(score),
                str(row.get("name", "")),
                str(row.get("headline", "")),
                str(row.get("company_name", "")),
                str(row.get("location", "")),
                str(row.get("linkedin_url", "")),
                str(row.get("email", "")),
                str(row.get("phone", "")),
                str(row.get("source", "")),
                mh_str,
                str(row.get("data_quality", "")),
                str(row.get("profile_text", ""))[:400],
            ])
            ri = ws.max_row
            for cell in ws[ri]:
                cell.fill = fill
                cell.alignment = wrap
            ws.row_dimensions[ri].height = 55

        ws2 = wb.create_sheet("Search Artefacts")
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 90
        bold = Font(bold=True)
        for label, val in [
            ("Generated On", str(date.today())),
            ("Boolean X-Ray String", boolean_str),
            ("JD Summary", jd_text[:800]),
        ]:
            ws2.append([label, val])
            ws2[f"A{ws2.max_row}"].font = bold
            ws2[f"B{ws2.max_row}"].alignment = Alignment(wrap_text=True)
            ws2.row_dimensions[ws2.max_row].height = 40

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception:
        buf = BytesIO()
        df.to_csv(buf, index=False)
        return buf.getvalue()

# ── Session state ─────────────────────────────────────────────────────────────

for _k in ["results_df", "source_stats", "boolean_str", "search_query", "jd_text"]:
    if _k not in st.session_state:
        st.session_state[_k] = None if _k in ["results_df", "source_stats"] else ""

# ── Load all secrets ──────────────────────────────────────────────────────────

EXA_KEY        = get_secret("EXA_API_KEY")
GOOGLE_KEY     = get_secret("GOOGLE_API_KEY")
GOOGLE_CSE_ID  = get_secret("GOOGLE_CSE_ID")
APOLLO_KEY     = get_secret("APOLLO_API_KEY")
PDL_KEY        = get_secret("PDL_API_KEY")
CRUSTDATA_KEY  = get_secret("CRUSTDATA_API_KEY")
FULLENRICH_KEY = get_secret("FULLENRICH_API_KEY")
LUSHA_KEY      = get_secret("LUSHA_API_KEY")

# (api_key_or_bool, display_name, secret_hint)
SEARCH_SOURCES = [
    (EXA_KEY,                          "Exa",        "EXA_API_KEY"),
    (GOOGLE_KEY and GOOGLE_CSE_ID,     "Google CSE", "GOOGLE_API_KEY + GOOGLE_CSE_ID"),
    (APOLLO_KEY,                       "Apollo",     "APOLLO_API_KEY"),
    (PDL_KEY,                          "PDL",        "PDL_API_KEY"),
    (CRUSTDATA_KEY,                    "Crustdata",  "CRUSTDATA_API_KEY"),
]
ENRICH_SOURCES = [
    (FULLENRICH_KEY, "FullEnrich", "FULLENRICH_API_KEY"),
    (LUSHA_KEY,      "Lusha",      "LUSHA_API_KEY"),
]
SOURCE_COLORS = {
    "Exa": "🔵", "Google CSE": "🔴", "Apollo": "🟠", "PDL": "🟣", "Crustdata": "🟤",
}

# ── Sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("## ⚙️ Settings")

    st.markdown("### 🔌 Search Sources")
    active_count = 0
    for key, name, hint in SEARCH_SOURCES:
        if key:
            st.caption(f"✅ **{SOURCE_COLORS.get(name, '⚫')} {name}** — active")
            active_count += 1
        else:
            st.caption(f"⬜ **{name}** — add `{hint}` to secrets")

    st.markdown("### 🧬 Enrichment")
    for key, name, hint in ENRICH_SOURCES:
        if key:
            st.caption(f"✅ **{name}** — active")
        else:
            st.caption(f"⬜ **{name}** — add `{hint}` to secrets")

    st.divider()
    num_results = st.slider("Results per source", 5, 50, 20)
    enrich_top_n = st.slider(
        "Enrich top N candidates", 0, 20, 0, 5,
        help="Run FullEnrich / Lusha on the top-ranked candidates after initial scoring. 0 = skip.",
    )

    st.divider()
    st.markdown("### 📖 About")
    st.caption(
        "Multi-source talent search: Exa · Google CSE · Apollo · PDL · Crustdata.  \n"
        "Enrichment: FullEnrich · Lusha.  \n"
        "Results are deduplicated and scored on a 100-point rubric."
    )

# ── Main UI ───────────────────────────────────────────────────────────────────

st.markdown("# 🔍 Talent Sourcing Portal")
st.caption("Paste a JD → search across multiple sources → deduplicate & score → export")

tab_search, tab_results, tab_string = st.tabs(["📋 Search", "📊 Results", "🔧 Search String"])

# ── TAB 1: Search ─────────────────────────────────────────────────────────────

with tab_search:
    col1, col2 = st.columns([3, 2])

    with col1:
        jd_text = st.text_area(
            "Job Description",
            height=300,
            placeholder="Paste the full JD here...",
            value=st.session_state.jd_text,
        )

    with col2:
        st.markdown("#### Quick Filters")
        location_filter = st.text_input("Location", value="Germany", placeholder="Germany, Munich...")
        seniority_opt = st.selectbox(
            "Seniority", ["Senior / Manager", "Staff / Principal", "Director / VP", "Any"]
        )
        industry_tags = st.multiselect(
            "Industries",
            ["industrial", "technology", "energy", "automotive", "fintech", "healthcare", "manufacturing"],
            default=["industrial", "technology", "energy"],
        )
        extra_must = st.text_input("Additional keyword", placeholder="e.g. MBA, fluent German")
        st.markdown("#### Refinement")
        refinement = st.text_input(
            "Refine search (natural language)",
            placeholder='e.g. "MBB alumni only" or "remove automotive"',
        )

    # Preview boolean string before running
    if jd_text.strip():
        preview_bool = build_boolean_string(jd_text, refinement)
        with st.expander("🔍 Preview & edit Boolean / X-Ray string before running", expanded=False):
            st.text_area(
                "Google X-Ray string (used by Google CSE source)",
                value=preview_bool, height=80, key="preview_bool",
            )
            st.caption(
                "Exa, Apollo, and PDL use a structured query derived from the JD text. "
                "The Boolean string above is used for the Google CSE LinkedIn X-Ray search."
            )

    if active_count == 0:
        st.warning("⚠️ No API keys configured. Add at least `EXA_API_KEY` to Streamlit secrets.")

    if st.button("🚀 Run Search", type="primary", use_container_width=True):
        if not jd_text.strip():
            st.error("Please paste a Job Description first.")
        elif active_count == 0:
            st.error("No API keys configured.")
        else:
            st.session_state.jd_text = jd_text
            jd_reqs = parse_jd(jd_text, location_filter, seniority_opt, industry_tags, extra_must)
            boolean_str = build_boolean_string(jd_text, refinement)
            st.session_state.boolean_str = boolean_str
            st.session_state.search_query = jd_reqs["query_text"]

            source_stats: dict = {}
            all_results: list = []
            active_sources = [(key, name) for key, name, _ in SEARCH_SOURCES if key]
            progress = st.progress(0, text="Starting multi-source search...")

            for i, (key, name) in enumerate(active_sources):
                progress.progress(i / len(active_sources), text=f"Searching {name}…")
                try:
                    if name == "Exa":
                        results = search_exa(jd_reqs["query_text"], EXA_KEY, num_results)
                    elif name == "Google CSE":
                        results = search_google_cse(jd_reqs["query_text"], GOOGLE_KEY, GOOGLE_CSE_ID, num_results)
                    elif name == "Apollo":
                        results = search_apollo(jd_reqs, APOLLO_KEY, num_results)
                    elif name == "PDL":
                        results = search_pdl(jd_reqs, PDL_KEY, num_results)
                    elif name == "Crustdata":
                        results = search_crustdata(jd_reqs, CRUSTDATA_KEY, num_results)
                    else:
                        results = []
                    source_stats[name] = len(results)
                    all_results.append(results)
                except Exception as e:
                    st.warning(f"⚠️ {name} error: {e}")
                    source_stats[name] = 0
                    all_results.append([])

            progress.progress(0.85, text="Merging & deduplicating…")
            merged = merge_candidates(all_results)

            # Enrichment: score first to pick top N, then enrich, then re-score rest
            if enrich_top_n > 0 and (FULLENRICH_KEY or LUSHA_KEY):
                for c in merged:
                    s = score_candidate(c["profile_text"], jd_reqs)
                    c["_pre_score"] = s["total"]
                merged.sort(key=lambda x: x.get("_pre_score", 0), reverse=True)
                top, rest = merged[:enrich_top_n], merged[enrich_top_n:]
                if FULLENRICH_KEY:
                    progress.progress(0.90, text="Enriching with FullEnrich…")
                    top = enrich_fullenrich(top, FULLENRICH_KEY)
                if LUSHA_KEY:
                    progress.progress(0.95, text="Enriching with Lusha…")
                    top = enrich_lusha(top, LUSHA_KEY)
                merged = top + rest

            progress.progress(0.97, text="Scoring…")
            candidates = []
            for c in merged:
                s = score_candidate(c["profile_text"], jd_reqs)
                c.update({
                    "must_have_detail": s["must_have_detail"],
                    "must_have":        s["must_have"],
                    "seniority":        s["seniority"],
                    "experience":       s["experience"],
                    "nice_to_have":     s["nice_to_have"],
                    "industry":         s["industry"],
                    "company_signal":   s["company_signal"],
                    "location_signal":  s["location_signal"],
                    "quality_penalty":  s["quality_penalty"],
                    "data_quality":     s["data_quality"],
                    "score":            s["total"],
                    "grade":            grade_label(s["total"]),
                })
                candidates.append(c)

            df = pd.DataFrame(candidates).sort_values("score", ascending=False).reset_index(drop=True)
            df["rank"] = df.index + 1
            st.session_state.results_df = df
            st.session_state.source_stats = source_stats
            progress.empty()

            stats_str = " · ".join(f"{k}: {v}" for k, v in source_stats.items())
            st.success(
                f"✅ **{len(df)} unique candidates** across {len(active_sources)} sources "
                f"({stats_str}). Switch to the **Results** tab."
            )

# ── TAB 2: Results ────────────────────────────────────────────────────────────

with tab_results:
    df = st.session_state.results_df

    if df is None:
        st.info("Run a search first to see results here.")
    else:
        # Per-source breakdown
        if st.session_state.source_stats:
            src_items = list(st.session_state.source_stats.items())
            cols = st.columns(len(src_items))
            for col, (src, cnt) in zip(cols, src_items):
                col.metric(f"{SOURCE_COLORS.get(src, '⚫')} {src}", cnt)
            st.divider()

        # Grade summary
        strong    = int((df["score"] >= 75).sum())
        potential = int(((df["score"] >= 50) & (df["score"] < 75)).sum())
        weak      = int((df["score"] < 50).sum())
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Total (deduplicated)", len(df))
        m2.metric("🟢 Strong Match", strong)
        m3.metric("🟡 Potential", potential)
        m4.metric("🔴 Weak Match", weak)
        st.divider()

        grade_filter = st.multiselect(
            "Filter by grade",
            ["🟢 Strong Match", "🟡 Potential", "🔴 Weak Match"],
            default=["🟢 Strong Match", "🟡 Potential"],
        )
        filtered_df = df[df["grade"].isin(grade_filter)] if grade_filter else df

        display_cols = ["rank", "score", "grade", "name", "headline", "company_name",
                        "location", "email", "phone", "source", "data_quality", "linkedin_url"]
        available = [c for c in display_cols if c in filtered_df.columns]

        st.dataframe(
            filtered_df[available].rename(columns={
                "rank": "Rank", "score": "Score", "grade": "Grade",
                "name": "Name", "headline": "Headline", "company_name": "Company",
                "location": "Location", "email": "Email", "phone": "Phone",
                "source": "Source", "data_quality": "Data Quality", "linkedin_url": "LinkedIn",
            }),
            use_container_width=True,
            height=420,
            column_config={
                "LinkedIn": st.column_config.LinkColumn("LinkedIn"),
                "Score": st.column_config.ProgressColumn("Score", min_value=0, max_value=100),
            },
            hide_index=True,
        )

        st.divider()
        st.markdown("#### 📋 Candidate Details")
        for _, row in filtered_df.head(10).iterrows():
            score = int(row.get("score", 0))
            with st.expander(
                f"{row.get('grade', '')} **{row.get('name', 'Unknown')}** — "
                f"{score}/100 | {row.get('headline', '')} | _{row.get('source', '')}_"
            ):
                c1, c2 = st.columns([2, 1])
                with c1:
                    li = row.get("linkedin_url", "")
                    if li:
                        st.markdown(f"**LinkedIn:** [{li}]({li})")
                    em = row.get("email", "")
                    if em:
                        st.markdown(f"**Email:** {em}")
                    ph = row.get("phone", "")
                    if ph:
                        st.markdown(f"**Phone:** {ph}")
                    st.markdown(
                        f"**Company:** {row.get('company_name', '—')}  "
                        f"| **Location:** {row.get('location', '—')}"
                    )
                    st.markdown("**Profile text:**")
                    st.caption(str(row.get("profile_text", ""))[:800])
                with c2:
                    st.markdown("**Score breakdown:**")
                    term_hits = row.get("must_have_detail", {})
                    if isinstance(term_hits, dict) and term_hits:
                        st.caption("JD terms found in profile:")
                        st.write(", ".join(term_hits.keys()))
                    st.metric("JD Term Match", f"{row.get('must_have', 0)}/40")
                    st.metric("Role Match", f"{row.get('role_match', 0)}/20")
                    st.metric("Seniority", f"{row.get('seniority', 0)}/15")
                    st.metric("Industry", f"{row.get('industry', 0)}/15")

        st.divider()
        st.markdown("#### 📥 Export")
        col_a, col_b = st.columns(2)
        with col_a:
            excel_bytes = build_excel_bytes(
                filtered_df, st.session_state.jd_text, st.session_state.boolean_str
            )
            st.download_button(
                "⬇️ Download Excel", data=excel_bytes,
                file_name=f"candidates_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, type="primary",
            )
        with col_b:
            st.download_button(
                "⬇️ Download CSV",
                data=filtered_df[available].to_csv(index=False),
                file_name=f"candidates_{date.today()}.csv",
                mime="text/csv", use_container_width=True,
            )

# ── TAB 3: Search String ──────────────────────────────────────────────────────

with tab_string:
    st.markdown("### Boolean / X-Ray Search String")
    st.caption("Used by Google CSE for LinkedIn X-Ray. Copy into Google for a manual search.")

    boolean_str = st.text_area(
        "Search String (editable)",
        value=st.session_state.boolean_str or "Run a search first to generate the string.",
        height=120,
        key="editable_boolean",
    )
    if boolean_str != st.session_state.boolean_str:
        st.session_state.boolean_str = boolean_str

    st.markdown("#### Last structured query (Exa / Apollo / PDL)")
    st.code(st.session_state.search_query or "(not yet run)", language="text")

    st.divider()
    if st.session_state.boolean_str:
        encoded = urllib.parse.quote(st.session_state.boolean_str)
        st.markdown(f"[🔗 Open X-Ray search in Google](https://www.google.com/search?q={encoded})")

    st.divider()
    st.markdown("#### Refinement Suggestions")
    st.markdown("""
| Intent | What to type in "Refine search" |
|---|---|
| Focus on a specific company | `only candidates from Siemens Healthineers` |
| Exclude an industry | `remove automotive` or `no pharma` |
| Require a language | `German native speakers only` |
| Narrow seniority | `only director or VP level` |
| Narrow location | `Munich only` or `Berlin and Hamburg` |
| Add a skill | `must have SAP experience` |
| Narrow background | `must have hospital or clinic experience` |
    """)
